import UCIQE
import UIQM
import cv2
import numpy as np
import os
import glob
import openpyxl
import shutil
import haze
import resize

def fetchimages(path ) :
    return[os.path.join(path,f)for f in os.listdir(path) if (f.endswith('.png') or f.endswith('.jpg') or f.endswith('.jpeg') or f.endswith('.webp')  or f.endswith('.jpeg') or f.endswith('.jfif') )]

def return_resized_images(file):
    image = cv2.imread(file)
    return(resize.resize(image))

def name_of_Image(source , a):
    
    d =source+ "\\"
    a = a.replace(d, '')
    return a

def paste_resized_images(image , c , nam):
    if(c == 1):
        cv2.imwrite("C:\\Users\\hp\\Desktop\\clear\\d_"+nam,image)
    elif(c == 0):
        cv2.imwrite("C:\\Users\\hp\\Desktop\\degraded\\d_"+nam,image)
    


def sep(source,degraded,clear,exel, row, m_uism, m_uicm, m_uiqm):

 #Insert the path of the exel sheet in which the values is to be pasted
 re=openpyxl.load_workbook(exel)
 sheet=re['Sheet1']

 filename = fetchimages(source)
 num=0
 i=len(filename)

 k=sheet.cell(row=1,column=1)
 k.value="NAME"
 k=sheet.cell(row=1,column=2)
 k.value="ucique"
 k=sheet.cell(row=1,column=3)
 k.value="uiqm"
 k=sheet.cell(row=1,column=4)
 k.value="uicm"
 k=sheet.cell(row=1,column=5)
 k.value="uism"
 k=sheet.cell(row=1,column=6)
 k.value="uiconm"
 k=sheet.cell(row=1,column=7)
 k.value="Haze"
 k=sheet.cell(row=1,column=8)
 k.value="chroma"
 k=sheet.cell(row=1,column=9)
 k.value="luminance"
 k=sheet.cell(row=1,column=10)
 k.value="saturation"
 k=sheet.cell(row=1,column=11)
 k.value="class"
 ro=row
 co=1

 while (num<i):
     co=1
     image = return_resized_images(filename[num])
     sharpness= UIQM.uism(image)
     colourfullness = UIQM.uicm(image)
     contrast = UIQM.uiconm(image)
     
     chroma = UCIQE . chroma(image)
     luminance = UCIQE . luminance(image)
     saturation = UCIQE . saturation (image)
     
     
     k=sheet.cell(row=ro,column=co)
     nam=name_of_Image(source, filename[num])
     if(row == 2):
         k.value="d_"+nam
     else :
         k.value = nam

     #paste UCIQE value in the excel sheet
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value=UCIQE.ucique(image)

     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value=(sharpness + colourfullness + contrast)
     m_uiqm = m_uiqm + (sharpness + colourfullness + contrast)
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value=colourfullness
     m_uicm = m_uicm + colourfullness
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value=sharpness
     m_uism = m_uism + sharpness

     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value=contrast
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value= haze.haze(filename[num])
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value= chroma
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value= luminance
     
     co+=1
     k=sheet.cell(row=ro,column=co)
     k.value= saturation
     
     if(sharpness > 5.00 and colourfullness>4.20  and ((sharpness + colourfullness + contrast)>9.5) and (haze.haze(filename[num]) > 950)):
         #shutil.copy(filename[num], clear)
         #paste_resized_images(image , 1 , nam[last:])
         k=sheet.cell(row=ro,column=11)
         k.value=1
     else: 
       #  shutil.copy(filename[num], degraded)
         #paste_resized_images(image , 0 , nam[last:])
         k=sheet.cell(row=ro,column=11)
         k.value=0
         
    

     num+=1
     ro+=1   
     print(num)
     re.save(exel)
     
    
 re.save(exel)
 
 return ro ,m_uism, m_uicm, m_uiqm